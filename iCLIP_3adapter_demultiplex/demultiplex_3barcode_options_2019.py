#!/usr/bin/python

"""
# iCLIP 3'barcode demultiplex
Script process and demultiplex the raw fastq file inserting the 3'barcode in between the 5' barcode.


## Main call usage:

    ./demultiplex_3barcode.py multiplexed_reads.fq.gz sample_annotation.xlsx
    ./python multiplexed_reads.fq.gz sample_annotation.xlsx synthetic_reads_my_script.xlsx -s 50 -e 0.2 -w 10 -O yes -r no                      # Keep all the indivual intermediate files subset 50 reads

    Optional arguments:
    -o --output_folder  Path of output folder. If not expecified all the output will go to the path where you call script
    -r --remove         Boolean to remove intermediate files (yes or no). Default "yes"
    -M --max_lenght     Max length of sequencing read. Default 200 (integer)
    -m --min_lenght     Min length of sequencing read. Default 1 (integer)
    -O --extra_output   Demultiplex and keep individual files with the 3 barcode (yes or no). Default "no"
    -s --subset         Subset number of reads (Integer). Default "no"
    -w, --word          Minimun word lenght for Illumina 3' adapter. Default 7 (integer)
    -e, --error         Maximun error tolerance on 5' barcode demultiplexing Default 0.1 (float)                                                # The level of error tolerance is adjusted by specifying a maximum error rate, which is 0.1 (=10%) by default. Use the -e option to set a different value. To determine the number of allowed errors, the maximum error rate is multiplied by the length of the match (and then rounded off). Good error rate is 0.1667 as weel


## Requirement:

demultiplex_3barcode.py needs "Make_histogram.R" Rscript to be present on the same folder of main script
Also need   cutadapt,
            R and ggplot2
            python and argparse, defaultdict, re, random, gzip


## Info:

- Adapters
5'barcode                   NNNNAAAAANNN
3'Illumina adapter                                                     3ILLUMINAADAPTER
3'barcode                                                         NNTGC
- Reads
Normal read                 NNNNAAAAANNN.........SEQUENCE..............3ILLUMINAADAPTER
New 3'barcode read          NNNNAAAAANNN.........SEQUENCE.........NNTGC3ILLUMINAADAPTER
- Output reads
Transformed upstream        NNNNNNTGCAAAAANNN....SEQUENCE..............3ILLUMINAADAPTER
Transformed downstream      NNNNAAAAATGCNNNNN....SEQUENCE..............3ILLUMINAADAPTER


## Process scheme and output demultiplexed file names:

1. Remove the 3' Illumina adapter:
    - Output:               Illumina_adapter3_removed.fq.gz                                             Reads removing the 3' Illumina adapter
                            Illumina_unknown_removed.fq.gz                                              Reads where 3' Illumina adapter can't be found
2. Demultiplex based on the 5' barcode:
    - Output:               demultiplexed_5barcode_{name}.fq
                            5barcode_not_found.fastq.gz
3. Demultiplex based on the 3'barcode and intercalate in between the 5'barcode
     - Output:              demultiplexed_5barcode_{name}_Illumina_3adapter_added.fq
                            demultiplexed_5barcode_{name}_demultiplexed_3barcode_{name}_Illumina_3adapter_added.fq
                            fastqfilename_cutadapt_Logfile.txt                                          Cutadapt Log of demultiplex and extraction
4. If "--extra_output" option is set to yes demultiplexed files will be generated
                            Demultiplexed_{name}_5barcode_{name}_3barcode_repositioned.fq               Extra output demultiplexed to individual files to use in other applications. Conserve the barcodes
                            Demultiplexed_{name}_remove5barcode_{name}_3barcode_repositioned.fq         Extra output demultiplexed to individual files to use in other applications. Barcodes removed and random nucleotides moved to the fastq header
5. Create histograms of read length
     - Output:              Histogram{5'barcode_name}.pdf                                               Histograms of read lengths containing 3'barcodes
6. Remove intermediate files and merge the final file:
    - Output:               L3modified_merged_{name_original_fastq_file}.fq.gz                          Final 3barcode demultiplexed and merged all the intermediate files
7. Modified sample_annotation file:
    - Output:               L3modified_{name_original_excel_file}.xlsx                  Final annotation file where 3'barcode have been intercalated in between the 5'barcode


"""

import os
import subprocess
from time import sleep
import argparse
import sys
from collections import defaultdict
import re
import random
import gzip


reload(sys)
sys.setdefaultencoding('utf8')


## Module loads to use on CAMP
print os.system("ml cutadapt/1.9.1-foss-2016b-Python-2.7.12")
# print os.system("ml Python/2.7.12-intel-2016b")
print os.system("ml use /camp/apps/eb/dev/modules/all")
print os.system("ml Python/2.7.12-foss-2016b")
print os.system("ml R-bundle-cnvkit/0.2-foss-2016b-R-foss-2016b-3.3.1-bioc-3.3-libX11-1.6.3")

from openpyxl import load_workbook
from openpyxl import Workbook


###################
#### Functions ####
###################

def get_script_path():
    return os.path.dirname(os.path.realpath(sys.argv[0]))

def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description="Demultiplex L3 barcode and insert the 3'barcode in between the 5' barcode.")
    parser.add_argument('multiplexed_reads', help="Multiplexed reads.")
    parser.add_argument('sample_annotation', help="iCount sample annotation file.")
    parser.add_argument('-o', '--output_folder', help="Path of output folder.", default=path)
    parser.add_argument('-r', '--remove_intermediates', help="Remove intermediate files(yes or no).", default='yes')
    parser.add_argument('-M', '--max_lenght', help="Max read lenght.", default=200)
    parser.add_argument('-m', '--min_lenght', help="Min read lenght.", default=1)
    parser.add_argument('-O', '--extra_output', help="Keep individual files with the 3' barcode (yes or no).", default='no')
    parser.add_argument('-s', '--subset', help="Subset number of reads (Integer).", default=0)
    parser.add_argument('-w', '--word', help="Minimun word lenght for Illumina 3' adapter.", default=7)
    parser.add_argument('-e', '--error', help="Maximun error tolerance on 5' barcode demultiplexing", default=0.1)                              # The level of error tolerance is adjusted by specifying a maximum error rate, which is 0.1 (=10%) by default. Use the -e option to set a different value. To determine the number of allowed errors, the maximum error rate is multiplied by the length of the match (and then rounded off). Good error rate is 0.1667 as weel

    return parser.parse_args()


def regex_L3_barcode(linker):
    '''
        Search and match L3 barcode
        It must be on the linker column with the configuration:         whatever-3'barcode
        Everything after the hyphen will be recognised as 3' barcode and extracted
        need to extract the first object
    '''
    m = re.search('(?<=-)\w+', linker)
    return (m)


def parse_annotation(excel_filename):
    '''
    :param      excel_filename
    :return:    all_samples         names list
                all_barcode5_set    set of unique 5' barcodes
                adapter3_set        set of unique 3' Illumina adapters
                demultiplexing_L3   dictionary of keys (5' barcodes) and values list of corresponding 3' adaptors
                md                  dictionary of dictionaies where key is sample name and key will be header of annotation and corresponding value     ## Not used at the moment but will be needed to reconstitute new annotation modifiying the 5' barcode architecture and remove the 3' barcode from the "linker" column
    '''

    # Initialise objects
    # Dictionary of samples. Key will be the sample name
    md = defaultdict(dict)

    # List of samples and adapters
    all_samples = []
    all_barcode5 = []
    all_adapter3 = []

    # Dictionary of list that contains witch L3 barcode are under each 5' barcode
    demultiplexing_L3 = {}

    # Load excel file and read the first sheet
    original_excel_wb = load_workbook(filename=excel_filename)
    original_excel_ws = original_excel_wb.worksheets[0]
    # Get the excel_header of the spread sheet to use as key of keys on the sample dictionary
    excel_header = [str(cell.value) for cell in original_excel_ws[1]]
    print "Excel header", excel_header

    # Iterate ove all rows on the annotation
    line_no=2
    for row in range(2, original_excel_ws.max_row + 1):

        # Retrieve the sample name on the first column
        Sample_name = str(original_excel_ws['A' + str(row)].value)

        # Create a list with all the element in the row
        row_list = [str(cell.value) for cell in original_excel_ws[row]]

        # Iterate over columns and populate dictionary where key is the sample name
        col = 0
        for element in row_list[0:(len(excel_header))]:
            md[Sample_name][excel_header[col]] = element
            col += 1

        # # Introduce a line number on the dictionary to keep the order when creating the modified excel file
        # md[Sample_name]['line_number'] = line_no
        # line_no+=1

        # Get all 5' barcode, 3' adapter and append to list
        barcode5 = md[Sample_name]["5' barcode"]
        barcode5 = barcode5.replace(",", "").replace("_0", "")

        adapter3 = md[Sample_name]["3' adapter"]
        adapter3 = adapter3.replace("_1", "").replace("_2", "").replace(",", "")

        linker = md[Sample_name]["linker"]

        all_samples.append(Sample_name)
        all_barcode5.append(barcode5)
        all_adapter3.append(adapter3)

        print "Sample  ", md[Sample_name]["Sample name"]
        print "5adapter", md[Sample_name]["5' barcode"]
        print "linker  ", md[Sample_name]["linker"], "\n"

        # Check if L3 barcode exist on the annotation column
        m = regex_L3_barcode(linker)

        if m:
            barcode3 = m.group(0)

            # Populate dictionary with the 5' barcode with the corresponding L3 barcodes
            if barcode5 in demultiplexing_L3.keys():                # if 5' adapter exist on dictionary create a list with all the L3 barcodes

                check_list = demultiplexing_L3[barcode5]

                if barcode3 in check_list:                          # Check if there are duplicated L3 barcodes under the same 5' barcode. Raise error and exit
                    print "\n!!!!! Warning !!!!!"
                    print "Barcode L3 %s is duplicated under the 5' barcode %s and sample cannot be demultiplexed \n\n" % (
                    barcode3, barcode5)
                    print "Program exiting"
                    sys.exit(1)

                else:
                    demultiplexing_L3[barcode5].append(barcode3)    # If there are several 3' barcodes under the same 5' barcode and they aren't duplicated include in the list

            else:                                                   # if 5' barcode doesn't exist on the dictionary create a new key with the 5' barcode and the corresponding L3 barcode
                demultiplexing_L3[barcode5] = [barcode3]

        # print "Disctionary of list to demultiplex", demultiplexing_L3

    print "Dictionary of list to demultiplex", demultiplexing_L3
    # Remove None if present and get unique Illumina L3 adapter
    while 'None' in all_adapter3:
        all_adapter3.remove('None')

    adapter3_set = set(all_adapter3)

    # Get unique 5' barcodes
    while 'None' in all_barcode5:
        all_barcode5.remove('None')

    all_barcode5_set = set(all_barcode5)

    return all_barcode5_set, adapter3_set, demultiplexing_L3       # return all_samples, all_barcode5_set, adapter3_set, demultiplexing_L3, md


def demultiplex_cutadapt(fastq_file_path, overlap, error, all_barcode5_set, adapter3_set):

    # Demultiplex 3' barcode
    # Actual script only acept one 3' Illumina adapter
    if len(adapter3_set) == 1:

        # Get the adapter3
        adapter3 = str(list(adapter3_set)[0])

        # Fastq file name
        fastq_file_name = os.path.basename(fastq_file_path)
        fastq_file_name = fastq_file_name.replace(".fastq.gz", "")
        fastq_file_name = os.path.join(path, fastq_file_name)

        # Remove cutadapt log file if exists
        print os.system("rm %s_cutadapt_Logfile.txt" % fastq_file_name)

        ## INCORPORATE ERROR RATE  cmd = ('cutadapt -e 0.1667 -a adapter3=%s -o %s_{name}.fq.gz %s 2>&1 | tee %s_cutadapt_Logfile.txt' % (adapter3, fastq_file_name,  fastq_file_path, fastq_file_name))
        cmd = ('cutadapt -a adapter3=%s --overlap %s -o Illumina_{name}_removed.fq.gz %s 2>&1 | tee -a %s_cutadapt_Logfile.txt' % (
        adapter3, overlap, fastq_file_path, fastq_file_name))
        print "CMD", cmd
        subprocess.check_call(cmd, shell=True)

    else:
        print "\n!!!!! Warning !!!!!"
        print "Program cannot handle two different 3' Illumina adapters\n\n"
        print "Program exiting"
        sys.exit(1)

    # Demultiplex 5' barcode
    # Create fasta file with all the 5' barcodes (no duplicates)
    adapter5_fasta_file_name = "adapter5.fasta"
    adapter5_fasta_file_name = os.path.join(path, adapter5_fasta_file_name)

    with open(adapter5_fasta_file_name, 'w') as fasta:
        for barcode in all_barcode5_set:
            if barcode == "None":
                continue
            else:
                fasta.write(">{0}\n^{0}\n".format(barcode))             # Create fasta file with all the 5' barcodes (anchored to the 5' end) to feed cutadapt

    cmd = ('cutadapt -g file:%s -e %s --no-indels --no-trim --untrimmed-o 5barcode_not_found.fastq.gz -o demultiplexed_5barcode_{name}.fq Illumina_adapter3_removed.fq.gz 2>&1 | tee -a %s_cutadapt_Logfile.txt' % (adapter5_fasta_file_name, error, fastq_file_name))  # Jernej insists not allowing insertion on the 5' barcode
    print "CMD", cmd
    subprocess.check_call(cmd, shell=True)

    return


# fastq parser class
class ParseFastQ(object):
    """Returns a read-by-read fastQ parser analogous to file.readline()"""

    def __init__(self, filePath, headerSymbols=['@', '+']):
        """Returns a read-by-read fastQ parser analogous to file.readline().
        Exmpl: parser.next()
        -OR-
        Its an iterator so you can do:
        for rec in parser:
            ... do something with rec ...
        rec is tuple: (seqHeader,seqStr,qualHeader,qualStr)
        """
        if filePath.endswith('.gz'):
            self._file = gzip.open(filePath)
        else:
            self._file = open(filePath, 'rU')
        self._currentLineNumber = 0
        self._hdSyms = headerSymbols

    def __iter__(self):
        return self

    def next(self):
        """Reads in next element, parses, and does minimal verification.
        Returns: tuple: (seqHeader,seqStr,qualHeader,qualStr)"""
        # ++++ Get Next Four Lines ++++
        elemList = []
        for i in range(4):
            line = self._file.readline()
            self._currentLineNumber += 1  ## increment file position
            if line:
                elemList.append(line.strip('\n'))
            else:
                elemList.append(None)

        # ++++ Check Lines For Expected Form ++++
        trues = [bool(x) for x in elemList].count(True)
        nones = elemList.count(None)
        # -- Check for acceptable end of file --
        if nones == 4:
            raise StopIteration
        # -- Make sure we got 4 full lines of data --
        assert trues == 4, \
            "** ERROR: It looks like I encountered a premature EOF or empty line.\n\
            Please check FastQ file near line number %s (plus or minus ~4 lines) and try again**" % (
                self._currentLineNumber)
        # -- Make sure we are in the correct "register" --
        assert elemList[0].startswith(self._hdSyms[0]), \
            "** ERROR: The 1st line in fastq element does not start with '%s'.\n\
            Please check FastQ file near line number %s (plus or minus ~4 lines) and try again**" % (
                self._hdSyms[0], self._currentLineNumber)
        assert elemList[2].startswith(self._hdSyms[1]), \
            "** ERROR: The 3rd line in fastq element does not start with '%s'.\n\
            Please check FastQ file near line number %s (plus or minus ~4 lines) and try again**" % (
                self._hdSyms[1], self._currentLineNumber)
        # -- Make sure the seq line and qual line have equal lengths --
        assert len(elemList[1]) == len(elemList[3]), "** ERROR: The length of Sequence data and Quality data of the last record aren't equal.\n\
               Please check FastQ file near line number %s (plus or minus ~4 lines) and try again**" % (
            self._currentLineNumber)

        # ++++ Return fatsQ data as tuple ++++
        return tuple(elemList)


# function to create list of list
def init_list_of_objects(size):
    ''' Function to create list of list'''
    list_of_objects = list()
    for i in range(0, size):
        list_of_objects.append(list())  # different object reference each time
    return list_of_objects


def move_L3(all_barcode5_set, demultiplexing_L3):

    # Synthetic 3' Illumina adapter
    adapter3 = "AGATCGGAAGAGCGGTTCAG"
    adapter3_qual = "FFFFFFFFFFFFFFFFFFFF"

    for barcode5 in all_barcode5_set:

        input_fastq = ("demultiplexed_5barcode_%s.fq" % (barcode5))
        counter = 0

        # On samples that was not needed demultiplexing the 3' barcode we will add a synthetic Illumina L3 adapter
        if barcode5 not in demultiplexing_L3:

            print ("\nAdding Illumina L3 adapter to %s file" % input_fastq)

            parser = ParseFastQ(input_fastq)  # optional arg: headerSymbols allows changing the header symbols
            for record in parser:

                # Define each element on a fastq file
                header = record[0]
                seq = record[1]
                header2 = record[2]
                qual = record[3]

                if len(seq) < Min_read_Lenght:
                    continue

                else:
                    counter += 1
                    # Print progress line every 10000 lines
                    if int(counter) % 10000 == 0:
                        print ".",

                    seq = seq + adapter3
                    qual = qual + adapter3_qual

                    output_fastq = ("demultiplexed_5barcode_%s_Illumina_3adapter_added.fq" % (barcode5))
                    output_fastq = os.path.join(path, output_fastq)

                    with open(output_fastq, 'a') as fasta:
                        fasta.write("%s\n%s\n%s\n%s\n" % (header, seq, header2, qual))

        # Samples that need to get demultiplexed based on the 3' barcode
        else:

            # for key in demultiplexing_L3:
            print ("\n")
            print ("5' barcode:     %s" % barcode5)
            print ("3' barcodes:    %s" % demultiplexing_L3[barcode5])

            bardode5_No_sym = barcode5
            barcodes = demultiplexing_L3[barcode5]

            print "Input fastq file", input_fastq
            input_fastq = os.path.join(path, input_fastq)

            # input_fastq_rename=input_fastq.replace(".fq", "")
            demultiplex_3barcodes_string = '_'.join(demultiplexing_L3[barcode5])
            out = ("demultiplexed_5barcode_%s_demultiplexed_3barcode_%s_Illumina_3adapter_added.fq" % (
            barcode5, demultiplex_3barcodes_string))
            fout = open(out, "wt")

            # Create list with the name of files to output demultiplex but with the barcode on
            final_files = []
            # Create list with the name of files to output demultiplex remove the barcode and move the rrandom barcode to teh header and unique ID for each read
            final_files_random = []

            # Create files to output reads removing the 5' barcode and the 3'L barcode
            for i in range(0, len(barcodes)):
                results_file_name = ("Demultiplexed_%s_5barcode_%s_3barcode_repositioned.fq" % (bardode5_No_sym, barcodes[i]))

                results_file = os.path.join(path, results_file_name)
                # print "RESULTS FILE", results_file
                final_files.append(results_file)
                final_files[i] = open(final_files[i], 'w')

                ## Create another set of files where we remove the barcode and move the random barcode to the heather
                results_file_name_random = ("Demultiplexed_%s_remove5barcode_%s_3barcode_repositioned.fq" % (bardode5_No_sym, barcodes[i]))

                results_file_random = os.path.join(path, results_file_name_random)
                # print "RESULTS FILE", results_file
                final_files_random.append(results_file_random)
                final_files_random[i] = open(final_files_random[i], 'w')

            counter_reads = 0
            counter_No_L3_barcode = 0
            counter_not_reach_L3 = 0
            counter_too_sort = 0

            # Create empty list for number of reads containin each L3 barcode
            list_barcode_presence = [0] * len(barcodes)
            # Create list of list with size of read for each L3 barcode
            list_read_lenght = init_list_of_objects(len(barcodes))

            print "Removing 3' adapter and inserting in between the 5' barcode\n"
            parser = ParseFastQ(input_fastq)
            for record in parser:

                # Initialise counter of total reads
                counter_reads += 1

                # Define each element on a fastq file
                header = record[0]
                seq = record[1]
                header2 = record[2]
                qual = record[3]

                # Discart reads that are sorter that ? because Illumina adapted could not be reached
                if len(seq) > Max_read_Lenght:
                    counter_not_reach_L3 += 1
                    continue

                # & discard reads that are sorter than 17 because they will only contain 5' adapter (12 nt). Allow 5 nt of seq
                elif len(seq) < Min_read_Lenght:
                    counter_too_sort += 1
                    continue

                # get the last 3 nucleotides of the sequence aka 3' barcode
                end = seq[-3:]

                # Check if the last 3 nt are on the barcodes pool
                if end in barcodes:

                    barcode_index = barcodes.index(end)

                    # insert downstream 5' barcode
                    # move the 3L barcode to the 5' in the sequence line and reposition inbetwen the 5' adapter
                    reposition_L3barcode = seq[-3:]
                    reposition_L3random = seq[-5:-3]
                    reposition = (reposition_L3barcode + reposition_L3random)

                    seq = seq[:-5]
                    seq = seq[:9] + reposition + seq[9:] # + adapter3                                                   # Not adding L3 Ilumina adapter anymore

                    # move the 3L barcode to the 5' in quality line and reposition in between the 5' adapter
                    reposition_qual_L3barcode = qual[-3:]
                    reposition_qual_L3random = seq[-5:-3]
                    reposition_qual = reposition_qual_L3barcode  + reposition_qual_L3random

                    qual = qual[:-5]
                    qual = qual[:9] + reposition_qual + qual[9:] #+ adapter3_qual                                       # Not adding L3 Ilumina adapter anymore

                    ###### Alternative to move the 3' barcode upstream the 5' experimental barcode
                    # insert Upstream 5' barcode
                    # ## move the 3L barcode to the 5' in the sequence line and reposition in betwen the 5' adapter
                    # reposition = seq[-5:]
                    # seq = seq[:-5]
                    # seq = seq[:4] + reposition + seq[4:] + adapter3

                    # ## move the 3L barcode to the 5' in quality line and repositinate in betwen the 5' adapter
                    # reposition_qual = qual[-5:]
                    # qual = qual[:-5]
                    # qual = qual[:4] + reposition_qual + qual[4:] + adapter3_qual
                    ######

                    # list of read count per 3' barcpde to make some stats
                    list_barcode_presence[barcode_index] = list_barcode_presence[barcode_index] + 1
                    list_read_lenght[barcode_index].append(len(seq))

                    # Add only the sequences that contains the L3 barcode
                    # Save fastq file with the 3' barcode incorporated in the midle of the 5' barcode
                    fout.write("%s\n%s\n%s\n%s\n" % (header, seq, header2, qual))

                    #################################
                    ##  When we need individual demultiplexed files with or without the barcode
                    #################################

                    # check if extra output is needed
                    if keep_extra_output == "yes":

                        # Individual output demultiplexed files

                        final_files[barcode_index].write("%s\n%s\n%s\n%s\n" % (header, seq, header2, qual))

                        ## Now remove all the adapter and move the random barcode to the read header and unique Identification for each read
                        random_barcode = seq[:4]+ seq[12:17]

                        seq_remove_L5L3=seq[17:]
                        qual_remove_L5L3=qual[17:]

                        header_add=("%s;%s;%s" % (header, random_barcode, list_barcode_presence[barcode_index]))

                        if len(seq_remove_L5L3)>=20:

                            final_files_random[barcode_index].write("%s\n%s\n%s\n%s\n" % (header_add, seq_remove_L5L3, header2, qual_remove_L5L3))

                    elif keep_extra_output == "no":
                        pass

                    else:
                        print('keep_extra_output argument should be yes or no')
                        sys.exit(1)

                else:
                    # print "Couldn't find L3 barcode in seq: \n" + "%s\n%s\n%s\n%s\n" % (header, seq, header2, qual)
                    counter_No_L3_barcode += 1

                # Add all the sequences doesn't mater if l3 barcode is present
                # Save fastq file with the 3' barcode incorporated in the midle of the 5' barcode
                # fout.write("%s\n%s\n%s\n%s\n" % (header, seq, header2, qual))

            # close multiplexed file
            fout.close()
            for i in range(0, len(barcodes)):
                final_files[i].close()
                final_files_random[i].close()

            print "3' barcodes", barcodes
            print "3' barc Count", list_barcode_presence
            # print list_read_lenght

            # Print some info and stats
            print ("Total number of reads:                      %s" % counter_reads)
            print ("Reads without L3 barcode:                   %s" % counter_No_L3_barcode)
            print ("Not reached L3 barcode:                     %s" % counter_not_reach_L3)
            print ("Sort reads that contains only 5' barcode:   %s" % counter_too_sort)

            # Create Histogram of read lenghts
            for index in range(len(barcodes)):
                barcode = barcodes[index]
                reads_with_barcode = int(list_barcode_presence[index])
                print ("\n\nReads with L3 %s barcode: %s" % (barcodes[index], int(list_barcode_presence[index])))

                read_lenght = list_read_lenght[index]

                print barcode
                print reads_with_barcode
                # print read_lenght

                if reads_with_barcode < 2:
                    print "Warning!! Not enought values to plot histogram"
                    continue

                else:

                    out_lenght = ("read_lenght_%s.txt" % (barcode))
                    results_file = os.path.join(path, out_lenght)
                    fout_reads_leng = open(results_file, "wt")

                    for item in read_lenght:
                        fout_reads_leng.write("%s " % item)

                    print os.system("chmod -R 777 %s" % results_file)
                    print "Results file", results_file

                    fout_reads_leng.close()
                    sleep(5)

                    cmd = ('Rscript "%s/Make_histogram.R" %s %s %s %s' % (script_path, results_file, bardode5_No_sym, barcode, path))
                    subprocess.call(cmd, shell=True)

                    # remove files of reads lenghts
                    cmd = ('rm -r %s' % (results_file))
                    subprocess.call(cmd, shell=True)

def merge_and_remove(remove_intermediates_files, fastq_file_path):
    '''
    Merge all the demultiplexed files and remove the intermediate files
    If argument remove_intermediates_files is "no" the intermediate files will not be removed.
    '''

    # Rename output fastq file
    fastq_file_name = os.path.basename(fastq_file_path)
    renamed_fastq_file_name = ("L3modified_merged_%s" % fastq_file_name)
    renamed_fastq_file_name = os.path.join(path, renamed_fastq_file_name)

    print os.system("cat *Illumina_3adapter_added.fq > temp_merge.fq")
    print os.system("gzip temp_merge.fq")
    print os.system("cat 5barcode_not_found.fastq.gz Illumina_unknown_removed.fq.gz temp_merge.fq.gz 2> /dev/null > %s" % renamed_fastq_file_name)              # Redirect error to NULL so if one of the files doesn't exist, concatenate still works
    print os.system("rm -r temp_merge.fq.gz")

    if remove_intermediates_files == "yes":
        # Remove all the intermediate files
        print os.system("rm -r demultiplexed_5barcode_* 5barcode_not_found.fastq.gz Illumina_adapter3_removed.fq.gz adapter5.fasta Rplots.pdf *_3barcode_repositioned.fq")

    elif remove_intermediates_files == "no":
        print("Not removing any intermediate files")
        print("Compressing intermediate files")
        print os.system("gzip *.fq")

    else:
        print('remove_intermediates_files argument should be yes or no')
        sys.exit(1)


def create_modified_annotation_file(excel_filename):

    # Load excel file and read the first sheet
    original_excel_wb = load_workbook(filename=excel_filename)
    original_excel_ws = original_excel_wb.worksheets[0]

    # Get the excel_header of the spread sheet
    excel_header = [str(cell.value) for cell in original_excel_ws[1]]
    # Limit header to the actual number of annotation columns
    excel_header = excel_header[0:27]
    #print "Excel header", excel_header

    # Load output excel file
    output_excel_filename = os.path.basename(excel_filename)
    output_excel_filename = ("L3modified_%s" % output_excel_filename)
    output_excel_filename = os.path.join(path, output_excel_filename)
    # Open output excel
    output_excel_wb = Workbook()
    output_excel_ws = output_excel_wb.create_sheet("Template", 0)
    # Populate header into the output annotation excel
    for i, excel_headerN in enumerate(excel_header):
        output_excel_ws.cell(row=1, column=i + 1).value = excel_headerN

    # Iterate over all rows on the annotation
    for row in range(2, original_excel_ws.max_row + 1):

        # Create a list with all the element in the row
        row_list = [str(cell.value) for cell in original_excel_ws[row]]
        # Limit list to the actual number of annotation columns
        row_list = row_list[0:27]

        if '-' in row_list[16]:                                                         # Recognise presence of 3'barcode on linker column
            m=regex_L3_barcode(row_list[16])
            barcode3 = m.group(0)
            row_list[11] = row_list[11].replace("_0,", "%s_0,NN" % barcode3)            # Intercalate 3'barcode in between 5'barcode

        for i, row_listN in enumerate(row_list):

            if row_listN == 'None':                                                     # Not populate empty cells
                continue
            else:
                output_excel_ws.cell(row=row, column=i+1).value = row_listN

    # Save output excel file
    output_excel_wb.save(output_excel_filename)


################
#### Inputs ####
################
"""Arguments passed when calling script"""

path = os.getcwd()
script_path = get_script_path()


args = parse_arguments()
fastq_file_path = args.multiplexed_reads
excel_filename = args.sample_annotation
path = args.output_folder
remove_intermediates_files = args.remove_intermediates
Max_read_Lenght = args.max_lenght
Min_read_Lenght = args.min_lenght
keep_extra_output = args.extra_output
subset = int(args.subset)
overlap = args.word
error = args.error

os.chdir(path)

################
"""Arguments passed when calling localy"""
# # # fastq file
# #fastq_file_path = "/camp/lab/ulej/inputs/babs/paulo.gameiro/mickael.escudero/PM17022/180216_K00102_0183_AHNLW2BBXX/fastq/GAM599A61-A87_S16_L005_R1_001.fastq.gz"
# fastq_file_path = "/Volumes/lab-ulej/working/pre-demux_for_iMaps/20181202_JU_lane35_PM18259/test_remove/merge_my_script.fq.gz"
# excel_filename = "/Volumes/lab-ulej/working/pre-demux_for_iMaps/20181202_JU_lane35_PM18259/test_remove/synthetic_reads_my_script.xlsx"
#
# # # fastq_file_path="/Volumes/lab-ulej/inputs/babs/christoph.sadee/mickael.escudero/Sadee_190917/171101_D00446_0261_ACBP73ANXX/fastq/SAD533A130-A162_S17_L005_R1_001.fastq.gz"
# # # Annotation excel file
# # # Parameters that need to exposed on the pipeline and probably changed after inspection of read length histograms
# # Max_read_Lenght = 150
# # Min_read_Lenght = 25
# #
# #path = ""
# remove_intermediates_files = "no"
# Max_read_Lenght = 150
# Min_read_Lenght = 20
# keep_extra_output = "yes"
# subset = 0
# overlap = 7
# error = 0.1


###################
####  Main     ####
###################

# Parse annotation
all_barcode5_set, adapter3_set, demultiplexing_L3 = parse_annotation(excel_filename)

# Modify annotation moving L3barcode in between the 5' barcode
create_modified_annotation_file(excel_filename)

# # If subset option is set fastq_file_path will be subset and name modified
if subset == 0:
    # No need for subset the data
    print "No need for subset the data"
    pass

elif subset > 0:

    # Fastq file name
    fastq_file_name = os.path.basename(fastq_file_path)
    fastq_file_name = fastq_file_name.replace(".fastq.gz", "")
    fastq_file_name = fastq_file_name.replace(".fq.gz", "")
    print ("\nSubsampling reads\n")
    new_fastq_file_name = ("%s_subset_%s.fastq.gz" % (fastq_file_name, subset))
    new_fastq_file_name = os.path.join(path, new_fastq_file_name)
    print "SUBSET FILE", new_fastq_file_name

    number_to_sample = subset

    total_records = 0
    cmd = ("cat %s | wc -l" % fastq_file_path)
    pipe = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE)
    total_records = pipe.communicate()
    total_records=  int(filter(str.isdigit, total_records[0]))
    print("sampling " + str(number_to_sample) + " out of " + str(total_records/4) + " records")

    ## Other randomisations
    # total_records = 0
    # parser = ParseFastQ(fastq_file_path)
    # for record in parser:
    #     total_records += 1
    # print("sampling " + str(number_to_sample) + " out of " + str(total_records) + " records PARSER")


    # total_records = 0
    # with gzip.open(fastq_file_path) as input:
    #     num_lines = sum([1 for line in input])
    # total_records = int(num_lines // 4)
    # print("sampling " + str(number_to_sample) + " out of " + str(total_records) + " records LINEs")


    # record_number = 0
    # with gzip.open(fastq_file_path) as input:
    #     with gzip.open(new_fastq_file_name, "w") as output:
    #         for line1 in input:
    #             line2 = input.next()
    #             line3 = input.next()
    #             line4 = input.next()
    #             if record_number in records_to_keep:
    #                     output.write(line1)
    #                     output.write(line2)
    #                     output.write(line3)
    #                     output.write(line4)
    #             record_number += 1
    #

    records_to_keep = set(random.sample(xrange((total_records/4) + 1), number_to_sample))

    record_number = 0
    with gzip.open(new_fastq_file_name, "w") as output:
        parser = ParseFastQ(fastq_file_path)
        for record in parser:
            if record_number in records_to_keep:

                # Define each element on a fastq file
                header = record[0]
                seq = record[1]
                header2 = record[2]
                qual = record[3]
                output.write("%s\n%s\n%s\n%s\n" % (header, seq, header2, qual))

            record_number += 1

    # Reasign the subset file to input for demultiplexing pipeline
    fastq_file_path = new_fastq_file_name

else:
    print "Must supply 'no' or integer with number of reads to keep for the subset option"
    print "Program exiting"
    sys.exit(1)



# Demultiplex using cutadapt
demultiplex_cutadapt(fastq_file_path, overlap, error, all_barcode5_set, adapter3_set)

# Move 3' barcode in between 5' barcode and plot histogram of read lengths
move_L3(all_barcode5_set, demultiplexing_L3)

# Merge all the files and remove intermediates files
merge_and_remove(remove_intermediates_files, fastq_file_path)

print("All done!")
sys.exit(1)